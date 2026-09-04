from __future__ import annotations

import logging
import uuid
from dataclasses import dataclass
from datetime import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import BondConfig
from .file_locks import cleanup_temporary, replace_with_retry, synchronized_path

logger = logging.getLogger(__name__)

SETTINGS_SHEET = "设置"
BONDS_SHEET = "监控列表"
ALERTS_SHEET = "提醒记录"
GUIDE_SHEET = "使用说明"
PARAMETERS_SHEET = "参数说明"
WORKBOOK_SCHEMA_VERSION = "1.4.1"
DEFAULT_UPDATE_MANIFEST_URL = (
    "https://github.com/hajiao/kzz-monitor/releases/latest/download/update-manifest.json"
)

BOND_HEADERS = [
    "启用", "转债代码", "名称", "卖出观察价", "回撤提醒%", "趋势窗口", "趋势最小跌幅",
    "建仓线", "加仓线", "重仓线", "当周评价", "当前价", "趋势", "近一年最高价",
    "监控峰值", "回撤%", "仓位区域", "强赎状态", "最近更新", "运行说明",
]

DEFAULT_SETTINGS: dict[str, Any] = {
    "轮询间隔秒": 60,
    "整轮间隔分钟": 60,
    "开盘时间": "09:30",
    "午间休市开始": "11:30",
    "午间休市结束": "13:00",
    "收盘时间": "15:00",
    "收盘后完整更新": True,
    "安道全文件": "",
    "安道全工作表": "",
    "桌面通知": True,
    "邮件通知": False,
    "邮件收件人": "",
    "SMTP服务器": "",
    "SMTP端口": 465,
    "SMTP用户名": "",
    "SMTP发件人": "",
    "SMTP使用SSL": True,
    "邮件冷却分钟": 30,
    "日志保留天数": 30,
    "更新清单地址": DEFAULT_UPDATE_MANIFEST_URL,
    "启动时检查更新": True,
    "更新检查间隔小时": 24,
    "自动安装更新": False,
    "工作簿结构版本": WORKBOOK_SCHEMA_VERSION,
}

GUIDE_ROWS = [
    ("KzzMonitor 快速使用指南", ""),
    ("1. 启动", "双击 KzzMonitor.exe；看到控制台后程序会自动开启轮询。"),
    ("2. 配置转债", "在“监控列表”填写启用、转债代码、卖出观察价、回撤%、趋势参数和三段线。"),
    ("轮询间隔区别", "轮询间隔秒=每查一支后的等待；整轮间隔分钟=两轮开始时间之差。整轮耗时超出时不额外等待。"),
    ("3. 保存关闭", "编辑完成后保存并关闭 Excel；程序写入工作簿时 Excel 必须关闭。"),
    ("4. 首次测试", "回到控制台点击“立即强制新一轮”，然后观察日志和监控列表的更新时间。"),
    ("5. 安道全", "控制台选择文件后会立即保存，并自动读取工作表名称供下拉选择；再点击刷新安道全。"),
    ("6. 邮件", "填写 SMTP 服务器、端口、用户名、发件人、授权码和收件人，保存后点击测试邮件。"),
    ("7. 后台", "关闭窗口时选“否”可隐藏到托盘继续运行；双击 KZZ 托盘图标恢复。"),
    ("8. 退出", "关闭窗口时选“是”，或在托盘菜单选择“退出 KzzMonitor”。"),
    ("9. 在线更新", "管理员配置更新清单地址后，点击“检查更新”；更新只替换程序和手册，不覆盖本工作簿、data 或 logs。"),
    ("自动更新策略", "默认启动时及每24小时检查；默认弹窗确认后安装。可开启自动安装，但程序会重启。"),
    ("10. 监控面板", "控制台“监控面板”页直接显示现价、趋势、峰值、回撤、三段线、评级、强赎和更新时间。"),
    ("监控颜色", "红=卖出；紫=重仓；橙=加仓；黄=建仓；粉=强赎；灰=无行情。深色背景使用白字保证可读。"),
    ("仅看提醒", "监控面板勾选后隐藏正常观察行；卖出锁定解除前会持续红色，不会随系统通知消失。"),
    ("11. 增删修改", "在监控面板选择行后修改表单并保存；同代码会更新原行，不会新增重复行。删除前会再次确认。"),
    ("快速查找回填", "输入部分代码或名称，从下拉结果选择后自动回填现有全部设置；代码/名称唯一匹配时按回车也可回填。"),
    ("12. 重复防呆", "轮询自动跳过相同代码的后续行并警告；“合并重复项”保留首行、补齐空值并删除重复行。"),
    ("状态：监控正常", "交易时段正在轮询。"),
    ("状态：等待开市", "开盘前、午休、收盘后或非交易日，无需处理。"),
    ("午休补跑", "若今天尚未完成过完整轮询，即使启动时处于午休，也会先补跑一轮再等待13:00。"),
    ("收盘断点续跑", "收盘最终轮询每完成一只就保存进度；在线更新或重启后跳过已完成代码，从下一只继续。"),
    ("提醒形式", "除邮件外是 Windows 通知中心/macOS 系统通知，不是阻塞弹窗；可在控制台测试。"),
    ("状态：异常重试中", "打开日志目录，查看最后一个错误；程序一般会自动重试。"),
    ("常见问题：无法写 Excel", "先保存并关闭正在打开的可转债监控.xlsx。"),
    ("Excel 待写队列", "工作簿被占用时，最新行情结果和每条提醒会存入数据库；关闭 Excel 后在下一次轮询自动补写。"),
    ("常见问题：无实时行情", "可能已退市或停止交易；核对最后交易日并将该行停用。"),
    ("安全提示", "SMTP 授权码不要写入 Excel；Windows 加密保存，macOS 存入钥匙串。"),
    ("完整手册", "同目录查看 KzzMonitor详细操作手册.html 或 PDF。"),
]

PARAMETER_ROWS = [
    ("参数", "推荐/示例", "详细说明"),
    ("轮询间隔秒（单只）", "60", "每查询完一支转债后的等待时间；最低 10 秒，绝对不是整轮或全表刷新间隔。"),
    ("整轮间隔分钟", "60", "按两轮开始时间计算；如果一轮本身耗时已超过该值，完成后不再额外等待。"),
    ("启用", "TRUE/FALSE", "TRUE 才进入轮询；退市或暂不关注的转债设为 FALSE。"),
    ("卖出观察价", "130", "当前波段实际达到此价格后，卖出回撤条件才会武装。"),
    ("回撤提醒%", "5", "(本波段峰值-当前价)/本波段峰值×100%，达到后还需趋势为下降。"),
    ("趋势窗口", "3–5", "使用最近多少次采样判断趋势；越大越平滑但越迟钝。"),
    ("趋势最小跌幅", "0.1", "窗口首尾的最小绝对变化；上涨和下降判断都使用该阈值。"),
    ("建仓线", "按评级填写", "现价小于等于此值时进入建仓区。"),
    ("加仓线", "≤建仓线", "现价小于等于此值时进入加仓区。"),
    ("重仓线", "≤加仓线", "现价小于等于此值时进入重仓区。通常建仓线≥加仓线≥重仓线。"),
    ("立即强制新一轮", "控制台按钮", "忽略是否开市，立即处理所有启用转债；仍遵守单只间隔。"),
    ("刷新安道全", "控制台按钮", "立即按代码覆盖当周评价、建仓线、加仓线和重仓线。"),
    ("SMTP SSL", "465=开；587=关", "关闭 SSL 时使用 STARTTLS；以邮箱服务商说明为准。"),
    ("监控峰值", "程序维护", "当前已突破观察价的合格波段峰值；低位新波段不会借用旧高峰，空闲时为0。"),
    ("短期波动过滤", "组合条件", "主要靠回撤提醒%，并同时受卖出观察价、趋势窗口、趋势最小跌幅及下跌次数共同过滤。"),
    ("均衡参数建议", "5% / 3–5 / 0.2", "回撤约5%、趋势窗口3–5、趋势最小跌幅0.2作为低频监控初始值，再按误报调整。"),
    ("近一年最高价", "程序维护", "最近 365 天历史最高价，当前主要用于展示。"),
    ("更新清单地址", "HTTPS/共享路径", "指向 update-manifest.json；更新只替换程序和手册，不覆盖用户 Excel、data 或 logs。"),
    ("更新检查间隔小时", "24", "程序保持运行时的自动检查周期；最低 1 小时。不会在每轮询前请求。"),
    ("自动安装更新", "FALSE", "FALSE 只提示确认；TRUE 下载校验后自动重启安装。建议保持 FALSE。"),
    ("Excel 待写队列", "自动", "Excel 被打开占用时，结果和提醒持久化到 data/monitor.db；下一次轮询自动重试。"),
    ("重复代码", "自动拦截", "新增相同代码会更新已有行；既有重复行本轮只处理第一行，可点击“合并重复项”。"),
    ("启用为空", "默认 TRUE", "只要代码有效，启用单元格为空时也按启用处理并写日志；明确 FALSE 才停用。"),
]

SETTING_DESCRIPTIONS = {
    "轮询间隔秒": "单只间隔：每查询完一支转债后的等待秒数，不是整轮间隔。",
    "整轮间隔分钟": "两轮开始时间的目标间隔；若一轮耗时已超过该值，则不额外等待。",
    "安道全文件": "安道全 Excel 的完整路径；留空则不导入。",
    "安道全工作表": "留空使用第一个工作表。",
    "邮件通知": "SMTP 密码不写入 Excel，使用环境变量，详见 README。",
    "邮件收件人": "多个地址用英文逗号分隔。",
    "SMTP服务器": "例如 smtp.qq.com；授权码在控制台中加密保存。",
    "更新清单地址": "可填 HTTPS、file:// 地址或本地/共享路径；留空不检查更新。",
    "更新检查间隔小时": "默认每 24 小时检查一次；最低 1 小时。",
    "自动安装更新": "建议 FALSE，避免交易时段未经确认重启。",
}


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用"}


def _as_enabled(value: Any, code: str) -> bool:
    """有代码但启用为空时默认启用；明确 FALSE 才停用。"""
    if value is None or str(value).strip() == "":
        logger.warning("转债 %s 的“启用”为空，按已启用处理；如需停用请明确填写 FALSE", code)
        return True
    return _as_bool(value)


def _float(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    return float(value)


def normalize_code(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    if text.startswith(("SH", "SZ", "BJ")):
        text = text[2:]
    return text.zfill(6) if text.isdigit() else text


@dataclass(slots=True)
class AppSettings:
    poll_interval_seconds: int = 60
    cycle_interval_minutes: int = 60
    open_time: time = time(9, 30)
    lunch_start: time = time(11, 30)
    lunch_end: time = time(13, 0)
    close_time: time = time(15, 0)
    final_cycle_after_close: bool = True
    adq_file: Path | None = None
    adq_sheet: str = ""
    desktop_notification: bool = True
    email_notification: bool = False
    email_recipients: tuple[str, ...] = ()
    smtp_host: str = ""
    smtp_port: int = 465
    smtp_user: str = ""
    smtp_from: str = ""
    smtp_ssl: bool = True
    email_cooldown_minutes: int = 30
    log_retention_days: int = 30
    update_manifest_url: str = ""
    check_updates_on_startup: bool = True
    update_check_interval_hours: int = 24
    auto_install_updates: bool = False


def create_workbook(path: Path, seed_codes: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    settings = wb.active
    settings.title = SETTINGS_SHEET
    settings.append(["配置项", "值", "说明"])
    for key, value in DEFAULT_SETTINGS.items():
        settings.append([key, value, SETTING_DESCRIPTIONS.get(key, "")])
    settings.column_dimensions["A"].width = 22
    settings.column_dimensions["B"].width = 52
    settings.column_dimensions["C"].width = 58

    bonds = wb.create_sheet(BONDS_SHEET)
    bonds.append(BOND_HEADERS)
    for code in seed_codes or ["127015", "110059"]:
        bonds.append([True, code, "", 130, 5, 3, 0.10])
    bonds.freeze_panes = "A2"
    bonds.auto_filter.ref = f"A1:T{max(2, bonds.max_row)}"
    widths = [8, 12, 16, 13, 12, 10, 15, 11, 11, 11, 12, 11, 10, 15, 12, 10, 12, 16, 20, 30]
    for index, width in enumerate(widths, 1):
        bonds.column_dimensions[chr(64 + index)].width = width

    alerts = wb.create_sheet(ALERTS_SHEET)
    alerts.append(["时间", "转债代码", "名称", "提醒类型", "价格", "消息"])
    alerts.freeze_panes = "A2"
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")
    _populate_help_sheets(wb)
    wb.save(path)


def _populate_help_sheets(wb: Any) -> None:
    for name in (GUIDE_SHEET, PARAMETERS_SHEET):
        if name in wb.sheetnames:
            del wb[name]
    guide = wb.create_sheet(GUIDE_SHEET, 0)
    for row in GUIDE_ROWS:
        guide.append(row)
    guide.merge_cells("A1:B1")
    guide["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    guide["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    guide["A1"].alignment = Alignment(horizontal="center")
    guide.column_dimensions["A"].width = 27
    guide.column_dimensions["B"].width = 100
    guide.freeze_panes = "A2"
    for row in range(2, guide.max_row + 1):
        guide.cell(row, 1).font = Font(bold=True, color="1F4E78")
        guide.cell(row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        guide.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[row].height = 34

    parameters = wb.create_sheet(PARAMETERS_SHEET, 1)
    for row in PARAMETER_ROWS:
        parameters.append(row)
    parameters.column_dimensions["A"].width = 24
    parameters.column_dimensions["B"].width = 20
    parameters.column_dimensions["C"].width = 90
    parameters.freeze_panes = "A2"
    parameters.auto_filter.ref = f"A1:C{parameters.max_row}"
    for cell in parameters[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    for row in range(2, parameters.max_row + 1):
        parameters.cell(row, 1).font = Font(bold=True, color="1F4E78")
        parameters.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
        parameters.row_dimensions[row].height = 32


@synchronized_path
def refresh_help_sheets(path: Path) -> None:
    wb = load_workbook(path)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.help{path.suffix}")
    try:
        _populate_help_sheets(wb)
        wb.save(temporary)
        wb.close()
        replace_with_retry(temporary, path)
    finally:
        wb.close()
        cleanup_temporary(temporary)


@synchronized_path
def migrate_workbook(path: Path) -> None:
    """原地补充新版配置和帮助页，不覆盖任何已有配置或监控数据。"""
    wb = load_workbook(path)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.migrating{path.suffix}")
    try:
        settings = wb[SETTINGS_SHEET]
        rows = {
            str(settings.cell(row, 1).value).strip(): row
            for row in range(2, settings.max_row + 1)
            if settings.cell(row, 1).value
        }
        changed = False
        for key, value in DEFAULT_SETTINGS.items():
            if key not in rows:
                settings.append([key, value, SETTING_DESCRIPTIONS.get(key, "")])
                rows[key] = settings.max_row
                changed = True
        schema_row = rows["工作簿结构版本"]
        if str(settings.cell(schema_row, 2).value or "") != WORKBOOK_SCHEMA_VERSION:
            settings.cell(schema_row, 2, WORKBOOK_SCHEMA_VERSION)
            changed = True
        if not changed:
            return
        _populate_help_sheets(wb)
        wb.save(temporary)
        wb.close()
        replace_with_retry(temporary, path)
    finally:
        wb.close()
        cleanup_temporary(temporary)


@synchronized_path
def load_configuration(path: Path) -> tuple[AppSettings, list[BondConfig]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        raw = {row[0]: row[1] for row in wb[SETTINGS_SHEET].iter_rows(min_row=2, values_only=True) if row[0]}
        parse_time = lambda value: time.fromisoformat(str(value))  # noqa: E731
        adq_text = str(raw.get("安道全文件") or "").strip()
        settings = AppSettings(
            poll_interval_seconds=max(10, int(raw.get("轮询间隔秒", 60))),
            cycle_interval_minutes=max(1, int(raw.get("整轮间隔分钟", 60))),
            open_time=parse_time(raw.get("开盘时间", "09:30")),
            lunch_start=parse_time(raw.get("午间休市开始", "11:30")),
            lunch_end=parse_time(raw.get("午间休市结束", "13:00")),
            close_time=parse_time(raw.get("收盘时间", "15:00")),
            final_cycle_after_close=_as_bool(raw.get("收盘后完整更新", True)),
            adq_file=Path(adq_text) if adq_text else None,
            adq_sheet=str(raw.get("安道全工作表") or ""),
            desktop_notification=_as_bool(raw.get("桌面通知", True)),
            email_notification=_as_bool(raw.get("邮件通知", False)),
            email_recipients=tuple(x.strip() for x in str(raw.get("邮件收件人") or "").split(",") if x.strip()),
            smtp_host=str(raw.get("SMTP服务器") or "").strip(),
            smtp_port=int(raw.get("SMTP端口") or 465),
            smtp_user=str(raw.get("SMTP用户名") or "").strip(),
            smtp_from=str(raw.get("SMTP发件人") or "").strip(),
            smtp_ssl=_as_bool(raw.get("SMTP使用SSL", True)),
            email_cooldown_minutes=max(1, int(raw.get("邮件冷却分钟", 30))),
            log_retention_days=max(1, int(raw.get("日志保留天数", 30))),
            update_manifest_url=str(raw.get("更新清单地址") or DEFAULT_UPDATE_MANIFEST_URL).strip(),
            check_updates_on_startup=_as_bool(raw.get("启动时检查更新", True)),
            update_check_interval_hours=max(1, int(raw.get("更新检查间隔小时", 24))),
            auto_install_updates=_as_bool(raw.get("自动安装更新", False)),
        )
        sheet = wb[BONDS_SHEET]
        headers = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}
        result: list[BondConfig] = []
        seen_codes: set[str] = set()
        for row in range(2, sheet.max_row + 1):
            get = lambda name: sheet.cell(row, headers[name]).value  # noqa: E731
            code = normalize_code(get("转债代码"))
            if not code or not _as_enabled(get("启用"), code):
                continue
            if code in seen_codes:
                logger.warning("监控列表存在重复转债 %s；本轮跳过重复行，可在监控面板一键合并", code)
                continue
            seen_codes.add(code)
            result.append(BondConfig(
                code=code,
                enabled=True,
                name=str(get("名称") or ""),
                sell_trigger_price=float(get("卖出观察价") or 130),
                sell_drawdown_pct=float(get("回撤提醒%") or 5),
                trend_window=max(2, int(get("趋势窗口") or 3)),
                trend_epsilon=float(get("趋势最小跌幅") or 0.1),
                build_line=_float(get("建仓线")),
                add_line=_float(get("加仓线")),
                heavy_line=_float(get("重仓线")),
                rating=str(get("当周评价") or ""),
            ))
        return settings, result
    finally:
        wb.close()


@synchronized_path
def update_settings(path: Path, values: dict[str, Any]) -> None:
    """更新控制面板允许编辑的设置项，不改动监控列表和运行结果。"""
    wb = load_workbook(path)
    temporary = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.settings{path.suffix}")
    try:
        sheet = wb[SETTINGS_SHEET]
        rows = {str(sheet.cell(row, 1).value).strip(): row for row in range(2, sheet.max_row + 1)}
        for key, value in values.items():
            if key not in rows:
                sheet.append([key, value, SETTING_DESCRIPTIONS.get(key, "")])
                rows[key] = sheet.max_row
            else:
                sheet.cell(rows[key], 2, value)
        wb.save(temporary)
        wb.close()
        replace_with_retry(temporary, path)
    finally:
        if getattr(wb, "_archive", None) is not None:
            wb.close()
        cleanup_temporary(temporary)
