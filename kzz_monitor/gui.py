from __future__ import annotations

import logging
import queue
import sys
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageDraw
import pystray

from . import __version__
from .clock import china_now
from .config import load_configuration, update_settings
from .excel_store import ExcelStore
from .monitor_view import classify_bond_row, match_bond_rows
from .notifier import Notifier
from .platform_utils import open_path
from .secrets import load_secret, save_secret
from .service import MonitorService
from .updater import UpdateInfo, check_for_update, stage_and_launch_update

logger = logging.getLogger(__name__)


class QueueLogHandler(logging.Handler):
    def __init__(self, messages: queue.Queue[str]) -> None:
        super().__init__()
        self.messages = messages
        self.setFormatter(logging.Formatter("%(asctime)s %(levelname)s | %(message)s", "%H:%M:%S"))

    def emit(self, record: logging.LogRecord) -> None:
        self.messages.put(self.format(record))


def tray_image() -> Image.Image:
    image = Image.new("RGBA", (64, 64), (30, 90, 165, 255))
    draw = ImageDraw.Draw(image)
    draw.ellipse((5, 5, 59, 59), fill=(45, 125, 210, 255), outline="white", width=3)
    draw.text((13, 18), "KZZ", fill="white", stroke_width=1)
    return image


class MonitorApp:
    def __init__(self, root: tk.Tk, workbook: Path, state_path: Path) -> None:
        self.root = root
        self.workbook = workbook
        self.state_path = state_path
        self.service: MonitorService | None = None
        self.worker: threading.Thread | None = None
        self.quitting = False
        self.messages: queue.Queue[str] = queue.Queue()
        self.ui_actions: queue.Queue[tuple[Any, tuple[Any, ...]]] = queue.Queue()
        self.log_handler = QueueLogHandler(self.messages)
        logging.getLogger().addHandler(self.log_handler)
        self.tray: pystray.Icon | None = None
        self.vars: dict[str, tk.Variable] = {}
        self.bond_vars: dict[str, tk.Variable] = {}
        self.excel_store = ExcelStore(workbook)
        self.monitor_refresh_job: str | None = None
        self.monitor_refresh_in_progress = False
        self.monitor_rows_by_code: dict[str, dict[str, Any]] = {}
        self.search_labels: dict[str, str] = {}
        self.only_alerts_var = tk.BooleanVar(value=False)
        self.update_check_job: str | None = None
        self.update_check_in_progress = False

        self.root.title("可转债监控控制台")
        self.root.geometry("980x720")
        self.root.minsize(860, 620)
        self.root.protocol("WM_DELETE_WINDOW", self.confirm_close)
        self._build_ui()
        self._load_settings()
        self._start_tray()
        self.root.after(200, self._drain_logs)
        self.root.after(100, self._drain_ui_actions)
        self.root.after(500, self.start_monitor)
        self.root.after(1000, self._refresh_status)
        self.root.after(1200, self.refresh_monitor_table)
        self.root.after(2500, self._startup_update_check)

    def _build_ui(self) -> None:
        style = ttk.Style()
        ui_font = "PingFang SC" if sys.platform == "darwin" else "Microsoft YaHei UI"
        style.configure("Title.TLabel", font=(ui_font, 15, "bold"))
        style.configure("Status.TLabel", font=(ui_font, 11, "bold"))
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True)
        container = ttk.Frame(notebook, padding=14)
        monitor_page = ttk.Frame(notebook, padding=10)
        notebook.add(container, text="运行控制")
        notebook.add(monitor_page, text="监控面板")
        self._build_monitor_page(monitor_page)

        top = ttk.Frame(container)
        top.pack(fill="x")
        ttk.Label(top, text=f"可转债监控控制台  v{__version__}", style="Title.TLabel").pack(side="left")
        self.status_label = ttk.Label(top, text="● 正在准备", foreground="#c27c00", style="Status.TLabel")
        self.status_label.pack(side="right")

        self.summary_label = ttk.Label(container, text="", padding=(0, 8))
        self.summary_label.pack(fill="x")

        settings_box = ttk.LabelFrame(container, text="常用设置（保存后下一轮生效）", padding=10)
        settings_box.pack(fill="x", pady=(4, 10))
        definitions = [
            ("轮询间隔秒", "单只查询间隔（秒）", "entry"),
            ("整轮间隔分钟", "整轮启动间隔（分钟）", "entry"),
            ("开盘时间", "开盘时间", "entry"),
            ("午间休市开始", "午休开始", "entry"),
            ("午间休市结束", "午休结束", "entry"),
            ("收盘时间", "收盘时间", "entry"),
            ("安道全文件", "安道全文件", "file"),
            ("安道全工作表", "安道全工作表", "sheet"),
            ("桌面通知", "桌面通知", "check"),
            ("邮件通知", "邮件通知", "check"),
            ("邮件收件人", "邮件收件人", "entry"),
            ("SMTP服务器", "SMTP 服务器", "entry"),
            ("SMTP端口", "SMTP 端口", "entry"),
            ("SMTP用户名", "SMTP 用户名", "entry"),
            ("SMTP发件人", "SMTP 发件人", "entry"),
            ("SMTP使用SSL", "SMTP 使用 SSL", "check"),
            ("SMTP授权码", "SMTP 授权码", "password"),
            ("更新清单地址", "更新清单地址", "entry"),
            ("启动时检查更新", "启动时检查更新", "check"),
            ("更新检查间隔小时", "更新检查间隔（小时）", "entry"),
            ("自动安装更新", "自动安装更新", "check"),
        ]
        for index, (key, label, kind) in enumerate(definitions):
            row, pair = divmod(index, 2)
            base = pair * 3
            ttk.Label(settings_box, text=label).grid(row=row, column=base, sticky="w", padx=(0, 6), pady=4)
            if kind == "check":
                variable: tk.Variable = tk.BooleanVar()
                ttk.Checkbutton(settings_box, variable=variable).grid(row=row, column=base + 1, sticky="w", pady=4)
            elif kind == "sheet":
                variable = tk.StringVar()
                self.adq_sheet_combo = ttk.Combobox(
                    settings_box, textvariable=variable, state="readonly", width=28
                )
                self.adq_sheet_combo.grid(row=row, column=base + 1, sticky="ew", pady=4)
                self.adq_sheet_combo.bind("<<ComboboxSelected>>", lambda _event: self.save_settings())
            else:
                variable = tk.StringVar()
                entry = ttk.Entry(settings_box, textvariable=variable, width=28 if kind != "file" else 42)
                if kind == "password":
                    entry.configure(show="●")
                entry.grid(row=row, column=base + 1, sticky="ew", pady=4)
                if kind == "file":
                    ttk.Button(settings_box, text="浏览", command=self._choose_adq).grid(row=row, column=base + 2, padx=(5, 8))
            self.vars[key] = variable
        settings_box.columnconfigure(1, weight=1)
        settings_box.columnconfigure(4, weight=1)

        actions = ttk.Frame(container)
        actions.pack(fill="x", pady=(0, 10))
        primary_actions = ttk.Frame(actions)
        primary_actions.pack(fill="x", pady=(0, 5))
        secondary_actions = ttk.Frame(actions)
        secondary_actions.pack(fill="x")
        self.start_button = ttk.Button(primary_actions, text="开启轮询", command=self.start_monitor)
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(primary_actions, text="结束轮询", command=self.stop_monitor)
        self.stop_button.pack(side="left", padx=6)
        ttk.Button(primary_actions, text="立即强制新一轮", command=self.force_cycle).pack(side="left", padx=6)
        ttk.Button(primary_actions, text="刷新安道全", command=self.refresh_adq).pack(side="left", padx=6)
        ttk.Button(primary_actions, text="保存设置", command=self.save_settings).pack(side="left", padx=6)
        ttk.Button(primary_actions, text="测试邮件", command=self.test_email).pack(side="left", padx=6)
        ttk.Button(primary_actions, text="测试桌面提醒", command=self.test_desktop).pack(side="left", padx=6)
        ttk.Button(secondary_actions, text="打开监控表", command=self.open_workbook).pack(side="left")
        ttk.Button(secondary_actions, text="操作手册", command=self.open_manual).pack(side="left", padx=6)
        ttk.Button(secondary_actions, text="检查更新", command=self.check_updates).pack(side="left", padx=6)
        ttk.Button(secondary_actions, text="打开日志目录", command=self.open_logs).pack(side="left", padx=6)
        ttk.Button(secondary_actions, text="隐藏到托盘", command=self.hide_window).pack(side="right")

        ttk.Label(container, text="运行日志").pack(anchor="w")
        log_frame = ttk.Frame(container)
        log_frame.pack(fill="both", expand=True, pady=(4, 0))
        self.log_text = tk.Text(log_frame, wrap="word", state="disabled", font=("Consolas", 9), bg="#101820", fg="#d7e3ef")
        scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

    def _build_monitor_page(self, parent: ttk.Frame) -> None:
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill="x", pady=(0, 7))
        legend = (
            ("卖出观察", "#B91C1C", "#FFFFFF"), ("重仓区", "#6D28D9", "#FFFFFF"),
            ("加仓区", "#F59E0B", "#1F2937"), ("建仓区", "#FDE68A", "#1F2937"),
            ("强赎关注", "#F9A8D4", "#4A044E"), ("无实时行情", "#D1D5DB", "#374151"),
        )
        ttk.Label(toolbar, text="颜色图例：").pack(side="left")
        for text, background, foreground in legend:
            tk.Label(toolbar, text=f" {text} ", bg=background, fg=foreground, padx=4).pack(side="left", padx=2)
        ttk.Checkbutton(
            toolbar, text="仅看提醒", variable=self.only_alerts_var,
            command=self.refresh_monitor_table,
        ).pack(side="right")

        columns = (
            "启用", "转债代码", "名称", "提醒状态", "当前价", "趋势", "近一年最高价", "监控峰值",
            "回撤%", "仓位区域", "建仓线", "加仓线", "重仓线", "当周评价", "最近更新",
            "强赎状态", "运行说明",
        )
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill="both", expand=True)
        self.monitor_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=16)
        widths = {"启用": 55, "转债代码": 85, "名称": 100, "提醒状态": 90, "当前价": 75, "趋势": 70,
                  "近一年最高价": 105, "监控峰值": 90, "回撤%": 70, "仓位区域": 80,
                  "建仓线": 75, "加仓线": 75, "重仓线": 75, "当周评价": 90, "最近更新": 145,
                  "强赎状态": 130, "运行说明": 260}
        for column in columns:
            self.monitor_tree.heading(column, text=column)
            self.monitor_tree.column(column, width=widths[column], anchor="center", stretch=False)
        vertical = ttk.Scrollbar(table_frame, orient="vertical", command=self.monitor_tree.yview)
        horizontal = ttk.Scrollbar(table_frame, orient="horizontal", command=self.monitor_tree.xview)
        self.monitor_tree.configure(yscrollcommand=vertical.set, xscrollcommand=horizontal.set)
        self.monitor_tree.tag_configure("sell", background="#B91C1C", foreground="#FFFFFF")
        self.monitor_tree.tag_configure("heavy", background="#6D28D9", foreground="#FFFFFF")
        self.monitor_tree.tag_configure("add", background="#F59E0B", foreground="#1F2937")
        self.monitor_tree.tag_configure("build", background="#FDE68A", foreground="#1F2937")
        self.monitor_tree.tag_configure("redeem", background="#F9A8D4", foreground="#4A044E")
        self.monitor_tree.tag_configure("unavailable", background="#D1D5DB", foreground="#374151")
        self.monitor_tree.tag_configure("normal-even", background="#FFFFFF", foreground="#17202A")
        self.monitor_tree.tag_configure("normal-odd", background="#EEF6FF", foreground="#17202A")
        self.monitor_tree.grid(row=0, column=0, sticky="nsew")
        vertical.grid(row=0, column=1, sticky="ns")
        horizontal.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.monitor_tree.bind("<<TreeviewSelect>>", self._load_selected_bond)

        editor = ttk.LabelFrame(parent, text="新增或修改监控转债", padding=8)
        editor.pack(fill="x", pady=(8, 0))
        search_line = ttk.Frame(editor)
        search_line.grid(row=0, column=0, columnspan=8, sticky="ew", pady=(0, 7))
        ttk.Label(search_line, text="快速查找（代码/名称）").pack(side="left", padx=(0, 6))
        self.bond_search_var = tk.StringVar()
        self.bond_search_combo = ttk.Combobox(
            search_line, textvariable=self.bond_search_var, state="normal", width=42
        )
        self.bond_search_combo.pack(side="left", fill="x", expand=True)
        self.bond_search_combo.bind("<KeyRelease>", self._update_bond_search_suggestions)
        self.bond_search_combo.bind("<<ComboboxSelected>>", self._select_bond_search_result)
        self.bond_search_combo.bind("<Return>", self._select_bond_search_result)
        ttk.Label(search_line, text="输入部分代码或名称，选择后自动回填").pack(side="left", padx=(8, 0))
        fields = (
            ("启用", "启用", "check"), ("转债代码", "转债代码", "entry"),
            ("名称", "名称", "entry"), ("卖出观察价", "卖出观察价", "entry"),
            ("回撤提醒%", "回撤%", "entry"), ("趋势窗口", "趋势窗口", "entry"),
            ("趋势最小跌幅", "最小变动", "entry"), ("建仓线", "建仓线", "entry"),
            ("加仓线", "加仓线", "entry"), ("重仓线", "重仓线", "entry"),
            ("当周评价", "当周评价", "entry"),
        )
        for index, (key, label, kind) in enumerate(fields):
            field_row, pair = divmod(index, 4)
            row = field_row + 1
            column = pair * 2
            ttk.Label(editor, text=label).grid(row=row, column=column, sticky="w", padx=(0, 4), pady=3)
            variable: tk.Variable = tk.BooleanVar(value=True) if kind == "check" else tk.StringVar()
            if kind == "check":
                ttk.Checkbutton(editor, variable=variable).grid(row=row, column=column + 1, sticky="w")
            else:
                ttk.Entry(editor, textvariable=variable, width=15).grid(row=row, column=column + 1, sticky="ew", padx=(0, 8))
            self.bond_vars[key] = variable
        for column in (1, 3, 5, 7):
            editor.columnconfigure(column, weight=1)
        buttons = ttk.Frame(editor)
        buttons.grid(row=4, column=0, columnspan=8, sticky="ew", pady=(7, 0))
        ttk.Button(buttons, text="新增/保存", command=self.save_bond).pack(side="left")
        ttk.Button(buttons, text="删除选中", command=self.delete_selected_bond).pack(side="left", padx=6)
        ttk.Button(buttons, text="清空表单", command=self.clear_bond_form).pack(side="left", padx=6)
        ttk.Button(buttons, text="刷新结果", command=self.refresh_monitor_table).pack(side="left", padx=6)
        ttk.Button(buttons, text="合并重复项", command=self.deduplicate_bonds).pack(side="left", padx=6)
        self.monitor_message = ttk.Label(buttons, text="", foreground="#9a6700")
        self.monitor_message.pack(side="right")
        self.clear_bond_form()

        for key in ("转债代码", "名称"):
            widget_name = None
            # Entry widgets are found by their Tk variable, avoiding a duplicate widget registry.
            for widget in editor.winfo_children():
                if isinstance(widget, ttk.Entry) and str(widget.cget("textvariable")) == str(self.bond_vars[key]):
                    widget_name = widget
                    break
            if widget_name is not None:
                widget_name.bind("<Return>", lambda _event, field=key: self._lookup_from_editor(field))
                widget_name.bind("<FocusOut>", lambda _event, field=key: self._lookup_from_editor(field))

    @staticmethod
    def _display_cell(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, float):
            return f"{value:.3f}".rstrip("0").rstrip(".")
        return str(value)

    def refresh_monitor_table(self) -> None:
        if self.monitor_refresh_job is not None:
            try:
                self.root.after_cancel(self.monitor_refresh_job)
            except tk.TclError:
                pass
            self.monitor_refresh_job = None
        if self.monitor_refresh_in_progress:
            return
        self.monitor_refresh_in_progress = True
        threading.Thread(
            target=self._load_monitor_rows_worker,
            daemon=True,
            name="KzzMonitorTableReader",
        ).start()

    def _load_monitor_rows_worker(self) -> None:
        try:
            rows = self.excel_store.list_bonds()
            self._post_ui(self._apply_monitor_rows, rows, None)
        except PermissionError:
            self._post_ui(self._apply_monitor_rows, [], "工作簿正在被外部 Excel 占用，稍后再刷新")
        except Exception as exc:
            logger.exception("刷新监控面板失败")
            self._post_ui(self._apply_monitor_rows, [], f"刷新失败：{exc}")

    def _apply_monitor_rows(self, rows: list[dict[str, Any]], error: str | None) -> None:
        self.monitor_refresh_in_progress = False
        if error:
            self.monitor_message.configure(text=error)
            self._schedule_monitor_refresh()
            return
        try:
            selected_code = str(self.bond_vars["转债代码"].get()) if self.bond_vars else ""
            self.monitor_rows_by_code = {
                str(row.get("转债代码") or ""): row for row in reversed(rows)
            }
            self._refresh_search_choices(rows)
            for item in self.monitor_tree.get_children():
                self.monitor_tree.delete(item)
            counts: dict[str, int] = {}
            alert_counts: dict[str, int] = {}
            columns = self.monitor_tree["columns"]
            for index, row in enumerate(rows):
                code = str(row.get("转债代码") or "")
                counts[code] = counts.get(code, 0) + 1
                sell_latched = False
                if self.service:
                    sell_latched = self.service.state.get_value(f"sell_latched:{code}", "0") == "1"
                visual = classify_bond_row(row, sell_latched)
                if self.only_alerts_var.get() and not visual.is_alert:
                    continue
                if visual.is_alert:
                    alert_counts[visual.label] = alert_counts.get(visual.label, 0) + 1
                values = [
                    visual.label if column == "提醒状态" else self._display_cell(row.get(column))
                    for column in columns
                ]
                tag = visual.key if visual.key != "normal" else f"normal-{'even' if index % 2 == 0 else 'odd'}"
                item = self.monitor_tree.insert("", "end", iid=f"row-{index}", values=values, tags=(tag,))
                if code == selected_code:
                    self.monitor_tree.selection_set(item)
            duplicates = {code: count for code, count in counts.items() if count > 1}
            if duplicates:
                detail = "，".join(f"{code}×{count}" for code, count in duplicates.items())
                self.monitor_message.configure(text=f"发现重复：{detail}；轮询只处理首行")
            else:
                alert_text = "，".join(f"{name}{count}" for name, count in alert_counts.items()) or "无提醒"
                self.monitor_message.configure(text=f"共 {len(rows)} 条，无重复；{alert_text}")
        except Exception as exc:
            logger.exception("刷新监控面板失败")
            self.monitor_message.configure(text=f"刷新失败：{exc}")
        finally:
            self._schedule_monitor_refresh()

    def _schedule_monitor_refresh(self) -> None:
        if not self.quitting:
            self.monitor_refresh_job = self.root.after(5000, self.refresh_monitor_table)

    def _post_ui(self, callback: Any, *args: Any) -> None:
        if not self.quitting:
            self.ui_actions.put((callback, args))

    def _drain_ui_actions(self) -> None:
        while True:
            try:
                callback, args = self.ui_actions.get_nowait()
            except queue.Empty:
                break
            try:
                callback(*args)
            except Exception:
                logger.exception("处理后台界面事件失败")
        if not self.quitting:
            self.root.after(100, self._drain_ui_actions)

    def _load_selected_bond(self, _event: object = None) -> None:
        selection = self.monitor_tree.selection()
        if not selection:
            return
        values = self.monitor_tree.item(selection[0], "values")
        columns = list(self.monitor_tree["columns"])
        row = dict(zip(columns, values))
        source = self.monitor_rows_by_code.get(str(row.get("转债代码") or ""))
        if source is None:
            return
        self._fill_bond_form(source)

    def _fill_bond_form(self, source: dict[str, Any]) -> None:
        for key, variable in self.bond_vars.items():
            value = source.get(key)
            if isinstance(variable, tk.BooleanVar):
                variable.set(bool(value))
            else:
                variable.set("" if value is None else value)
        code = str(source.get("转债代码") or "")
        name = str(source.get("名称") or "")
        self.bond_search_var.set(f"{code} | {name}".rstrip(" |"))

    def _refresh_search_choices(self, rows: list[dict[str, Any]]) -> None:
        self.search_labels = {}
        for row in rows:
            code = str(row.get("转债代码") or "").strip()
            name = str(row.get("名称") or "").strip()
            label = f"{code} | {name}".rstrip(" |")
            self.search_labels[label] = code
        self.bond_search_combo.configure(values=list(self.search_labels))

    def _update_bond_search_suggestions(self, _event: object = None) -> None:
        query = self.bond_search_var.get()
        rows = list(self.monitor_rows_by_code.values())
        matches = match_bond_rows(query, rows)
        labels = [
            f"{str(row.get('转债代码') or '').strip()} | {str(row.get('名称') or '').strip()}".rstrip(" |")
            for row in matches
        ]
        self.bond_search_combo.configure(values=labels)
        if query.strip() and labels:
            self.bond_search_combo.event_generate("<Down>")

    def _select_bond_search_result(self, _event: object = None) -> None:
        query = self.bond_search_var.get().strip()
        code = self.search_labels.get(query)
        source = self.monitor_rows_by_code.get(code or "")
        if source is None:
            matches = match_bond_rows(query, list(self.monitor_rows_by_code.values()))
            if len(matches) == 1:
                source = matches[0]
            elif len(matches) > 1:
                self.monitor_message.configure(text=f"找到 {len(matches)} 条匹配，请从下拉列表选择")
                return
        if source is not None:
            self._fill_bond_form(source)
            self._select_tree_code(str(source.get("转债代码") or ""))

    def _lookup_from_editor(self, field: str) -> None:
        query = str(self.bond_vars[field].get()).strip()
        if not query:
            return
        matches = match_bond_rows(query, list(self.monitor_rows_by_code.values()))
        if len(matches) == 1:
            self._fill_bond_form(matches[0])
            self._select_tree_code(str(matches[0].get("转债代码") or ""))
            self.monitor_message.configure(text="已匹配现有转债并回填当前设置")
        elif len(matches) > 1:
            self.bond_search_var.set(query)
            self._update_bond_search_suggestions()
            self.monitor_message.configure(text=f"输入匹配到 {len(matches)} 只，请使用快速查找下拉选择")

    def _select_tree_code(self, code: str) -> None:
        columns = list(self.monitor_tree["columns"])
        code_index = columns.index("转债代码")
        for item in self.monitor_tree.get_children():
            values = self.monitor_tree.item(item, "values")
            if len(values) > code_index and str(values[code_index]) == code:
                self.monitor_tree.selection_set(item)
                self.monitor_tree.focus(item)
                self.monitor_tree.see(item)
                return

    def clear_bond_form(self) -> None:
        defaults: dict[str, Any] = {
            "启用": True, "转债代码": "", "名称": "", "卖出观察价": 130,
            "回撤提醒%": 5, "趋势窗口": 3, "趋势最小跌幅": 0.1,
            "建仓线": "", "加仓线": "", "重仓线": "", "当周评价": "",
        }
        for key, value in defaults.items():
            self.bond_vars[key].set(value)
        if hasattr(self, "bond_search_var"):
            self.bond_search_var.set("")
        if hasattr(self, "monitor_tree"):
            self.monitor_tree.selection_remove(self.monitor_tree.selection())

    def save_bond(self) -> None:
        try:
            values = {key: variable.get() for key, variable in self.bond_vars.items()}
            code, created = self.excel_store.upsert_bond(values)
            action = "新增" if created else "更新已有记录（未创建重复行）"
            self._append_log(f"监控转债 {code}：{action}")
            self.refresh_monitor_table()
        except PermissionError:
            messagebox.showerror("无法保存", "工作簿正在外部 Excel 中打开，请关闭后重试。")
        except Exception as exc:
            messagebox.showerror("参数有误", str(exc))

    def delete_selected_bond(self) -> None:
        code = str(self.bond_vars["转债代码"].get()).strip()
        if not code:
            messagebox.showinfo("删除监控", "请先在结果表中选择一只转债。")
            return
        if not messagebox.askyesno("删除监控", f"确定删除转债 {code} 的所有重复行和监控配置吗？"):
            return
        try:
            if self.excel_store.delete_bond(code):
                if self.service:
                    self.service.state.delete_pending_for_code(code)
                self._append_log(f"已删除监控转债：{code}")
                self.clear_bond_form()
                self.refresh_monitor_table()
        except PermissionError:
            messagebox.showerror("无法删除", "工作簿正在外部 Excel 中打开，请关闭后重试。")

    def deduplicate_bonds(self) -> None:
        try:
            duplicates = self.excel_store.deduplicate_bonds()
            if duplicates:
                detail = "，".join(f"{code} 删除 {count} 行" for code, count in duplicates.items())
                messagebox.showinfo("重复项已合并", detail)
                self._append_log(f"监控列表重复项已合并：{detail}")
            else:
                messagebox.showinfo("重复检查", "没有发现重复转债代码。")
            self.refresh_monitor_table()
        except PermissionError:
            messagebox.showerror("无法合并", "工作簿正在外部 Excel 中打开，请关闭后重试。")

    def _load_settings(self) -> None:
        settings, bonds = load_configuration(self.workbook)
        values: dict[str, Any] = {
            "轮询间隔秒": settings.poll_interval_seconds,
            "整轮间隔分钟": settings.cycle_interval_minutes,
            "开盘时间": settings.open_time.strftime("%H:%M"),
            "午间休市开始": settings.lunch_start.strftime("%H:%M"),
            "午间休市结束": settings.lunch_end.strftime("%H:%M"),
            "收盘时间": settings.close_time.strftime("%H:%M"),
            "安道全文件": str(settings.adq_file or ""),
            "安道全工作表": settings.adq_sheet,
            "桌面通知": settings.desktop_notification,
            "邮件通知": settings.email_notification,
            "邮件收件人": ",".join(settings.email_recipients),
            "SMTP服务器": settings.smtp_host,
            "SMTP端口": settings.smtp_port,
            "SMTP用户名": settings.smtp_user,
            "SMTP发件人": settings.smtp_from,
            "SMTP使用SSL": settings.smtp_ssl,
            "SMTP授权码": load_secret(self.state_path.parent / "smtp_secret.bin"),
            "更新清单地址": settings.update_manifest_url,
            "启动时检查更新": settings.check_updates_on_startup,
            "更新检查间隔小时": settings.update_check_interval_hours,
            "自动安装更新": settings.auto_install_updates,
        }
        for key, value in values.items():
            self.vars[key].set(value)
        try:
            self._load_adq_sheet_options(settings.adq_file, settings.adq_sheet)
        except Exception:
            self.adq_sheet_combo.configure(values=[])
            logger.exception("读取已配置的安道全工作表列表失败")
        self.summary_label.configure(text=f"配置文件：{self.workbook}    已启用转债：{len(bonds)} 只")

    def save_settings(self) -> bool:
        try:
            interval = max(10, int(str(self.vars["轮询间隔秒"].get()).strip()))
            cycle_interval = max(1, int(str(self.vars["整轮间隔分钟"].get()).strip()))
            update_interval = max(1, int(str(self.vars["更新检查间隔小时"].get()).strip()))
            for key in ("开盘时间", "午间休市开始", "午间休市结束", "收盘时间"):
                datetime.strptime(str(self.vars[key].get()).strip(), "%H:%M")
            values = {key: variable.get() for key, variable in self.vars.items()}
            values["轮询间隔秒"] = interval
            values["整轮间隔分钟"] = cycle_interval
            values["更新检查间隔小时"] = update_interval
            password = str(values.pop("SMTP授权码", ""))
            values["SMTP端口"] = int(str(values["SMTP端口"]).strip())
            update_settings(self.workbook, values)
            save_secret(self.state_path.parent / "smtp_secret.bin", password)
            self._load_settings()
            self._schedule_periodic_update_check(update_interval)
            self._append_log("设置已保存；后台将在下一轮重新读取。")
            return True
        except PermissionError:
            messagebox.showerror("无法保存", "监控工作簿正在 Excel 中打开。请先保存并关闭 Excel。")
            return False
        except Exception as exc:
            messagebox.showerror("设置有误", str(exc))
            return False

    def _choose_adq(self) -> None:
        selected = filedialog.askopenfilename(title="选择安道全 Excel", filetypes=[("Excel 工作簿", "*.xlsx *.xlsm"), ("所有文件", "*.*")])
        if selected:
            self.vars["安道全文件"].set(selected)
            try:
                self._load_adq_sheet_options(Path(selected), "")
            except Exception as exc:
                messagebox.showerror("无法读取工作表", str(exc))
                return
            if self.save_settings():
                self._append_log(f"已选择并保存安道全文件：{selected}")

    def _load_adq_sheet_options(self, path: Path | None, preferred: str) -> None:
        if not hasattr(self, "adq_sheet_combo"):
            return
        names: list[str] = []
        if path and path.exists():
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                names = list(workbook.sheetnames)
            finally:
                workbook.close()
        self.adq_sheet_combo.configure(values=names)
        selected = preferred if preferred in names else (names[0] if names else "")
        self.vars["安道全工作表"].set(selected)

    def start_monitor(self) -> None:
        if self.worker and self.worker.is_alive():
            self.show_window()
            return
        self.service = MonitorService(self.workbook, self.state_path, status_callback=self._service_status)
        self.worker = threading.Thread(target=self._run_service, name="KzzMonitorWorker", daemon=True)
        self.worker.start()
        self.status_label.configure(text="● 监控运行中", foreground="#14833b")
        self._update_tray_menu()

    def _service_status(self, state: str, message: str) -> None:
        self._post_ui(self._apply_service_status, state, message)

    def _apply_service_status(self, state: str, message: str) -> None:
        labels = {
            "checking": ("● 正在检查", "#c27c00"),
            "running": ("● 监控正常", "#14833b"),
            "waiting": ("● 非交易时段", "#2563a8"),
            "error": ("● 异常重试中", "#b42318"),
            "stopped": ("● 已停止", "#b42318"),
        }
        text, color = labels.get(state, ("● 运行中", "#14833b"))
        self.status_label.configure(text=text, foreground=color)
        self._append_log(f"状态：{message}")
        if self.tray:
            self.tray.title = f"可转债监控 - {message}"

    def _run_service(self) -> None:
        assert self.service is not None
        try:
            self.service.run(install_signal_handlers=False)
        except Exception:
            logger.exception("监控线程异常退出")
        finally:
            if not self.quitting:
                self._post_ui(self._mark_stopped)

    def stop_monitor(self) -> None:
        if self.service:
            self.service.stop()
        self.status_label.configure(text="● 正在停止", foreground="#c27c00")

    def force_cycle(self) -> None:
        if not self.worker or not self.worker.is_alive():
            self.start_monitor()
        if self.service:
            self.service.request_force_cycle()
            self._append_log("已请求立即开始新一轮完整轮询（仍按设置的单只间隔执行）。")

    def refresh_adq(self) -> None:
        if not self.worker or not self.worker.is_alive():
            self.start_monitor()
        if self.service:
            self.service.request_adq_refresh()
            self._append_log("已请求立即重新导入安道全评级和三段线。")

    def test_email(self) -> None:
        if not self.save_settings():
            return
        try:
            settings, _ = load_configuration(self.workbook)
            success = Notifier(settings, self.state_path.parent / "smtp_secret.bin").send_email_test()
            if success:
                messagebox.showinfo("测试成功", "测试邮件已经发出，请检查收件箱和垃圾邮件目录。")
            else:
                messagebox.showerror("测试失败", "邮件未发送。请检查 SMTP 设置，并查看运行日志。")
        except Exception as exc:
            logger.exception("测试邮件失败")
            messagebox.showerror("测试失败", str(exc))

    def test_desktop(self) -> None:
        try:
            settings, _ = load_configuration(self.workbook)
            Notifier(settings, self.state_path.parent / "smtp_secret.bin").send_desktop_test()
            self._append_log("已发送 Windows 通知中心/系统通知测试。")
        except Exception as exc:
            logger.exception("测试桌面提醒失败")
            messagebox.showerror("测试失败", str(exc))

    def _mark_stopped(self) -> None:
        self.service = None
        self.worker = None
        self.status_label.configure(text="● 已停止", foreground="#b42318")
        self._update_tray_menu()

    def _refresh_status(self) -> None:
        running = bool(self.worker and self.worker.is_alive())
        now = china_now().strftime("%Y-%m-%d %H:%M:%S")
        current = "运行中" if running else "已停止"
        pending = 0
        last_cycle = "尚无记录"
        if self.service and running:
            try:
                pending = self.service.state.pending_excel_write_count()
                stored = self.service.state.get_value("last_cycle_completed_at")
                if stored:
                    last_cycle = datetime.fromisoformat(stored).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pending = 0
        self.root.title(f"可转债监控控制台 - {current}")
        base_text = self.summary_label.cget("text").split("    北京时间")[0]
        self.summary_label.configure(
            text=f"{base_text}    北京时间：{now}    上次完整轮询：{last_cycle}    Excel待写：{pending} 条"
        )
        if not self.quitting:
            self.root.after(1000, self._refresh_status)

    def _drain_logs(self) -> None:
        while True:
            try:
                self._append_log(self.messages.get_nowait())
            except queue.Empty:
                break
        if not self.quitting:
            self.root.after(200, self._drain_logs)

    def _append_log(self, text: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def open_workbook(self) -> None:
        open_path(self.workbook)

    def open_logs(self) -> None:
        folder = self.state_path.parent.parent / "logs"
        folder.mkdir(parents=True, exist_ok=True)
        open_path(folder)

    def open_manual(self) -> None:
        base = self.state_path.parent.parent
        for name in ("KzzMonitor详细操作手册.html", "KzzMonitor详细操作手册.pdf", "README.md"):
            manual = base / name
            if manual.exists():
                open_path(manual)
                return
        messagebox.showinfo("操作手册", "当前目录没有操作手册，请重新解压完整便携包。")

    def _startup_update_check(self) -> None:
        try:
            settings, _ = load_configuration(self.workbook)
            if settings.check_updates_on_startup and settings.update_manifest_url:
                self.check_updates(silent_when_current=True, automatic=True)
            self._schedule_periodic_update_check(settings.update_check_interval_hours)
        except Exception:
            logger.exception("启动时检查更新失败")

    def _schedule_periodic_update_check(self, hours: int) -> None:
        if self.update_check_job is not None:
            try:
                self.root.after_cancel(self.update_check_job)
            except tk.TclError:
                pass
        self.update_check_job = self.root.after(
            max(1, hours) * 60 * 60 * 1000, self._periodic_update_check
        )

    def _periodic_update_check(self) -> None:
        self.update_check_job = None
        try:
            settings, _ = load_configuration(self.workbook)
            if settings.update_manifest_url:
                self.check_updates(silent_when_current=True, automatic=True)
            self._schedule_periodic_update_check(settings.update_check_interval_hours)
        except Exception:
            logger.exception("周期检查更新失败")
            self._schedule_periodic_update_check(1)

    def check_updates(self, silent_when_current: bool = False, automatic: bool = False) -> None:
        try:
            if self.update_check_in_progress:
                if not silent_when_current:
                    messagebox.showinfo("检查更新", "更新检查正在进行中。")
                return
            location = str(self.vars["更新清单地址"].get()).strip()
            if not location:
                if not silent_when_current:
                    messagebox.showinfo("检查更新", "请先填写“更新清单地址”并保存设置。")
                return
            self._append_log("正在检查程序更新……")
            self.update_check_in_progress = True
            threading.Thread(
                target=self._check_updates_worker,
                args=(location, silent_when_current, automatic),
                daemon=True,
                name="KzzUpdateChecker",
            ).start()
        except Exception as exc:
            messagebox.showerror("检查更新失败", str(exc))

    def _check_updates_worker(self, location: str, silent_when_current: bool, automatic: bool) -> None:
        try:
            info = check_for_update(location)
            self._post_ui(self._show_update_result, info, silent_when_current, automatic)
        except Exception as exc:
            logger.exception("检查更新失败")
            if not silent_when_current:
                self._post_ui(messagebox.showerror, "检查更新失败", str(exc))
            self._post_ui(self._finish_update_check)

    def _finish_update_check(self) -> None:
        self.update_check_in_progress = False

    def _show_update_result(self, info: UpdateInfo | None, silent_when_current: bool, automatic: bool) -> None:
        self._finish_update_check()
        if info is None:
            self._append_log(f"当前已是最新版本 v{__version__}。")
            if not silent_when_current:
                messagebox.showinfo("检查更新", f"当前已是最新版本 v{__version__}。")
            return
        notes = f"\n\n更新说明：\n{info.notes}" if info.notes else ""
        settings, _ = load_configuration(self.workbook)
        auto_install = automatic and settings.auto_install_updates
        install = auto_install or messagebox.askyesno(
                "发现新版本",
                f"当前版本：v{__version__}\n新版本：v{info.version}{notes}\n\n"
                "是否立即下载并安装？程序会自动退出、替换并重启；Excel、data 和 logs 不会被覆盖。",
            )
        if not install:
            return
        if auto_install:
            self._append_log(f"发现 v{info.version}，已启用自动安装。")
        self._append_log(f"正在下载并校验 v{info.version} 更新包……")
        threading.Thread(
            target=self._install_update_worker,
            args=(info,),
            daemon=True,
            name="KzzUpdateInstaller",
        ).start()

    def _install_update_worker(self, info: UpdateInfo) -> None:
        try:
            stage_and_launch_update(info, self.state_path.parent.parent)
            self._post_ui(self._exit_main)
        except Exception as exc:
            logger.exception("安装更新失败")
            self._post_ui(messagebox.showerror, "安装更新失败", str(exc))

    def hide_window(self) -> None:
        self.root.withdraw()
        if self.tray:
            try:
                self.tray.notify("程序仍在后台运行。双击托盘图标可重新打开控制台。", "可转债监控")
            except Exception:
                pass

    def confirm_close(self) -> None:
        choice = messagebox.askyesnocancel(
            "关闭 KzzMonitor",
            "是否彻底退出 KzzMonitor？\n\n"
            "“是”：停止轮询并退出程序\n"
            "“否”：隐藏到后台，继续运行\n"
            "“取消”：返回控制台",
            icon="question",
        )
        if choice is True:
            self._exit_main()
        elif choice is False:
            self.hide_window()

    def show_window(self, *_: object) -> None:
        self._post_ui(self._show_window_main)

    def _show_window_main(self) -> None:
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()

    def exit_app(self, *_: object) -> None:
        self._post_ui(self._exit_main)

    def _exit_main(self) -> None:
        self.quitting = True
        if self.service:
            self.service.stop()
        if self.tray:
            self.tray.stop()
        logging.getLogger().removeHandler(self.log_handler)
        self.root.destroy()

    def _start_tray(self) -> None:
        menu = pystray.Menu(
            pystray.MenuItem("显示控制台", self.show_window, default=True),
            pystray.MenuItem("启动监控", lambda *_: self._post_ui(self.start_monitor)),
            pystray.MenuItem("停止监控", lambda *_: self._post_ui(self.stop_monitor)),
            pystray.MenuItem("立即强制新一轮", lambda *_: self._post_ui(self.force_cycle)),
            pystray.MenuItem("刷新安道全", lambda *_: self._post_ui(self.refresh_adq)),
            pystray.MenuItem("检查更新", lambda *_: self._post_ui(self.check_updates)),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("退出 KzzMonitor", self.exit_app),
        )
        tray_options: dict[str, Any] = {}
        if sys.platform == "darwin":
            from AppKit import NSApplication

            tray_options["darwin_nsapplication"] = NSApplication.sharedApplication()
        self.tray = pystray.Icon(
            "KzzMonitor", tray_image(), "可转债监控 - 正在启动", menu, **tray_options
        )
        self.tray.run_detached()

    def _update_tray_menu(self) -> None:
        if self.tray:
            running = bool(self.worker and self.worker.is_alive())
            self.tray.title = f"可转债监控 - {'运行中' if running else '已停止'}"
            self.tray.update_menu()


def run_gui(workbook: Path, state_path: Path) -> None:
    root = tk.Tk()
    MonitorApp(root, workbook, state_path)
    root.mainloop()
