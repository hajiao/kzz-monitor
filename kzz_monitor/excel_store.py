from __future__ import annotations

import logging
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .config import ALERTS_SHEET, BOND_HEADERS, BONDS_SHEET, normalize_code
from .file_locks import workbook_lock
from .models import Evaluation, Quote

logger = logging.getLogger(__name__)
HEADER_INDEX = {name: index + 1 for index, name in enumerate(BOND_HEADERS)}
def synchronized(method: Any) -> Any:
    def wrapper(self: "ExcelStore", *args: Any, **kwargs: Any) -> Any:
        with self._lock:
            return method(self, *args, **kwargs)
    return wrapper


def _value(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


class ExcelStore:
    def __init__(self, path: Path) -> None:
        self.path = path
        self._lock = workbook_lock(path)

    @synchronized
    def import_adq(self, source: Path, sheet_name: str = "") -> int:
        if not source.exists():
            raise FileNotFoundError(f"安道全文件不存在: {source}")
        source_wb = load_workbook(source, data_only=True, read_only=True)
        try:
            sheet = source_wb[sheet_name] if sheet_name else source_wb.worksheets[0]
            header_row, headers = self._find_headers(sheet)
            records: dict[str, dict[str, Any]] = {}
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                code = normalize_code(row[headers["代码"]])
                if not code:
                    continue
                records[code] = {
                    "建仓线": _value(row[headers["建仓线"]]),
                    "加仓线": _value(row[headers["加仓线"]]),
                    "重仓线": _value(row[headers["重仓线"]]),
                    "当周评价": row[headers["当周评价"]] or "",
                }
        finally:
            source_wb.close()

        wb = load_workbook(self.path)
        try:
            target = wb[BONDS_SHEET]
            target_headers = {str(cell.value).strip(): cell.column for cell in target[1] if cell.value}
            count = 0
            for row_number in range(2, target.max_row + 1):
                code = normalize_code(target.cell(row_number, target_headers["转债代码"]).value)
                record = records.get(code)
                if not record:
                    continue
                for name, value in record.items():
                    target.cell(row_number, target_headers[name], value)
                count += 1
            self._safe_save(wb)
            return count
        finally:
            wb.close()

    @staticmethod
    def _find_headers(sheet: Any) -> tuple[int, dict[str, int]]:
        required = {"代码", "建仓线", "加仓线", "重仓线"}
        rating_aliases = ("当周评价", "当周评级", "评级")
        for row_number in range(1, min(sheet.max_row, 50) + 1):
            values = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[row_number]]
            positions = {value: index for index, value in enumerate(values) if value}
            rating = next((name for name in rating_aliases if name in positions), None)
            if required.issubset(positions) and rating:
                positions["当周评价"] = positions[rating]
                return row_number, positions
        raise ValueError("安道全工作表前 50 行未找到：代码、建仓线、加仓线、重仓线、当周评价/评级")

    @synchronized
    def update_result(self, quote: Quote, evaluation: Evaluation, explanation: str = "") -> None:
        wb = load_workbook(self.path)
        try:
            sheet = wb[BONDS_SHEET]
            row_number = self._find_bond_row(sheet, quote.code)
            values = {
                "名称": quote.name,
                "当前价": quote.price,
                "趋势": evaluation.trend.value,
                "近一年最高价": evaluation.one_year_high,
                "监控峰值": evaluation.monitored_peak,
                "回撤%": evaluation.drawdown_pct,
                "仓位区域": evaluation.zone,
                "强赎状态": quote.redeem_status,
                "最近更新": quote.timestamp.replace(tzinfo=None),
                "运行说明": explanation,
            }
            for name, value in values.items():
                sheet.cell(row_number, HEADER_INDEX[name], value)
            color = "FFC7CE" if evaluation.sell_alert else "FFF2CC" if evaluation.zone != "观察" else "FFFFFF"
            sheet.cell(row_number, HEADER_INDEX["当前价"]).fill = PatternFill("solid", fgColor=color)
            self._safe_save(wb)
        finally:
            wb.close()

    @synchronized
    def append_alert(self, quote: Quote, alert_type: str, message: str) -> None:
        wb = load_workbook(self.path)
        try:
            sheet = wb[ALERTS_SHEET]
            sheet.append([quote.timestamp.replace(tzinfo=None), quote.code, quote.name, alert_type, quote.price, message])
            self._safe_save(wb)
        finally:
            wb.close()

    @synchronized
    def mark_unavailable(self, code: str, checked_at: datetime, explanation: str) -> None:
        wb = load_workbook(self.path)
        try:
            sheet = wb[BONDS_SHEET]
            row_number = self._find_bond_row(sheet, code)
            values = {
                "当前价": None,
                "趋势": "无实时行情",
                "强赎状态": "可能已退市/停止交易",
                "最近更新": checked_at.replace(tzinfo=None),
                "运行说明": explanation,
            }
            for name, value in values.items():
                sheet.cell(row_number, HEADER_INDEX[name], value)
            sheet.cell(row_number, HEADER_INDEX["当前价"]).fill = PatternFill("solid", fgColor="D9D9D9")
            self._safe_save(wb)
        finally:
            wb.close()

    @synchronized
    def list_bonds(self) -> list[dict[str, Any]]:
        wb = load_workbook(self.path, data_only=True, read_only=True)
        try:
            sheet = wb[BONDS_SHEET]
            headers = [str(cell.value or "").strip() for cell in sheet[1]]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                for row in sheet.iter_rows(min_row=2, values_only=True)
                if normalize_code(row[HEADER_INDEX["转债代码"] - 1])
            ]
        finally:
            wb.close()

    @synchronized
    def upsert_bond(self, values: dict[str, Any]) -> tuple[str, bool]:
        code = normalize_code(values.get("转债代码"))
        if not code.isdigit() or len(code) != 6:
            raise ValueError("转债代码必须是六位数字")
        build = _value(values.get("建仓线"))
        add = _value(values.get("加仓线"))
        heavy = _value(values.get("重仓线"))
        sell_trigger = _value(values.get("卖出观察价"))
        drawdown = _value(values.get("回撤提醒%"))
        epsilon = _value(values.get("趋势最小跌幅"))
        try:
            trend_window = int(values.get("趋势窗口") or 3)
        except (TypeError, ValueError) as exc:
            raise ValueError("趋势窗口必须是大于等于 2 的整数") from exc
        if sell_trigger is None or sell_trigger <= 0:
            raise ValueError("卖出观察价必须是正数")
        if drawdown is None or not 0 < drawdown < 100:
            raise ValueError("回撤提醒%必须大于 0 且小于 100")
        if trend_window < 2:
            raise ValueError("趋势窗口必须大于等于 2")
        if epsilon is None or epsilon < 0:
            raise ValueError("趋势最小跌幅必须是非负数")
        configured = [item for item in (build, add, heavy) if item is not None]
        if len(configured) >= 2 and configured != sorted(configured, reverse=True):
            raise ValueError("三段线应满足：建仓线 ≥ 加仓线 ≥ 重仓线")
        wb = load_workbook(self.path)
        try:
            sheet = wb[BONDS_SHEET]
            matches = [
                row for row in range(2, sheet.max_row + 1)
                if normalize_code(sheet.cell(row, HEADER_INDEX["转债代码"]).value) == code
            ]
            created = not matches
            row_number = matches[0] if matches else sheet.max_row + 1
            values = dict(values)
            values["转债代码"] = code
            values.update({
                "卖出观察价": sell_trigger, "回撤提醒%": drawdown,
                "趋势窗口": trend_window, "趋势最小跌幅": epsilon,
                "建仓线": build, "加仓线": add, "重仓线": heavy,
            })
            for name in BOND_HEADERS[:11]:
                if name in values:
                    sheet.cell(row_number, HEADER_INDEX[name], values[name])
            sheet.auto_filter.ref = f"A1:T{sheet.max_row}"
            self._safe_save(wb)
            return code, created
        finally:
            wb.close()

    @synchronized
    def delete_bond(self, code: str) -> bool:
        normalized = normalize_code(code)
        wb = load_workbook(self.path)
        try:
            sheet = wb[BONDS_SHEET]
            rows = [
                row for row in range(2, sheet.max_row + 1)
                if normalize_code(sheet.cell(row, HEADER_INDEX["转债代码"]).value) == normalized
            ]
            for row in reversed(rows):
                sheet.delete_rows(row)
            if rows:
                sheet.auto_filter.ref = f"A1:T{max(2, sheet.max_row)}"
                self._safe_save(wb)
            return bool(rows)
        finally:
            wb.close()

    @synchronized
    def deduplicate_bonds(self) -> dict[str, int]:
        wb = load_workbook(self.path)
        try:
            sheet = wb[BONDS_SHEET]
            first_rows: dict[str, int] = {}
            duplicate_rows: list[int] = []
            duplicates: dict[str, int] = {}
            for row in range(2, sheet.max_row + 1):
                code = normalize_code(sheet.cell(row, HEADER_INDEX["转债代码"]).value)
                if not code:
                    continue
                if code not in first_rows:
                    first_rows[code] = row
                    sheet.cell(row, HEADER_INDEX["转债代码"], code)
                    continue
                target = first_rows[code]
                duplicates[code] = duplicates.get(code, 0) + 1
                for column in range(1, len(BOND_HEADERS) + 1):
                    if sheet.cell(target, column).value in (None, "") and sheet.cell(row, column).value not in (None, ""):
                        sheet.cell(target, column, sheet.cell(row, column).value)
                duplicate_rows.append(row)
            for row in reversed(duplicate_rows):
                sheet.delete_rows(row)
            if duplicate_rows:
                sheet.auto_filter.ref = f"A1:T{max(2, sheet.max_row)}"
                self._safe_save(wb)
            return duplicates
        finally:
            wb.close()

    @staticmethod
    def _find_bond_row(sheet: Any, code: str) -> int:
        code_column = HEADER_INDEX["转债代码"]
        for row in range(2, sheet.max_row + 1):
            if normalize_code(sheet.cell(row, code_column).value) == code:
                return row
        raise KeyError(f"监控列表中没有 {code}")

    def _safe_save(self, wb: Any) -> None:
        temporary = self.path.with_name(f".{self.path.stem}.writing{self.path.suffix}")
        try:
            wb.save(temporary)
            for attempt in range(5):
                try:
                    os.replace(temporary, self.path)
                    break
                except PermissionError:
                    if attempt == 4:
                        raise
                    time.sleep(0.2 * (attempt + 1))
        except PermissionError as exc:
            try:
                temporary.unlink(missing_ok=True)
            except PermissionError:
                pass
            raise PermissionError(f"无法更新 {self.path.name}，请关闭正在打开的 Excel 文件") from exc
        finally:
            try:
                temporary.unlink(missing_ok=True)
            except PermissionError:
                pass
