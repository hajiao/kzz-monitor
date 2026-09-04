from __future__ import annotations

import argparse
import logging
import os
import shutil
import sys
from contextlib import contextmanager
from logging.handlers import TimedRotatingFileHandler
from pathlib import Path
from typing import Generator

from openpyxl import load_workbook

from .config import create_workbook, load_configuration, migrate_workbook, normalize_code
from .excel_store import ExcelStore
from .notifier import Notifier
from .platform_utils import bundled_resource, instance_lock_path, show_already_running, user_data_dir
from .service import MonitorService


def application_dir() -> Path:
    return user_data_dir()


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    template = bundled_resource("可转债监控.xlsx")
    if template and template.resolve() != path.resolve():
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template, path)
    else:
        create_workbook(path, None)


def setup_logging(base: Path) -> None:
    log_dir = base / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(name)s | %(message)s")
    file_handler = TimedRotatingFileHandler(
        log_dir / "kzz_monitor.log", when="midnight", backupCount=30, encoding="utf-8"
    )
    file_handler.setFormatter(formatter)
    console = logging.StreamHandler()
    console.setFormatter(formatter)
    logging.basicConfig(level=logging.INFO, handlers=[file_handler, console])


def legacy_codes(path: Path) -> list[str]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        for sheet in wb.worksheets:
            for header_row in range(1, min(sheet.max_row, 30) + 1):
                for cell in sheet[header_row]:
                    if str(cell.value or "").strip() != "代码":
                        continue
                    result: list[str] = []
                    for row in range(header_row + 1, sheet.max_row + 1):
                        code = normalize_code(sheet.cell(row, cell.column).value)
                        if code.isdigit() and len(code) == 6:
                            result.append(code)
                    if result:
                        return list(dict.fromkeys(result))
        return []
    finally:
        wb.close()


@contextmanager
def single_instance(base: Path) -> Generator[None, None, None]:
    lock_path = instance_lock_path()
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    handle = lock_path.open("a+")
    try:
        if os.name == "nt":
            import msvcrt

            try:
                handle.seek(0)
                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            except OSError as exc:
                raise RuntimeError("已有一个可转债监控实例正在运行") from exc
        else:
            import fcntl

            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except OSError as exc:
                raise RuntimeError("已有一个可转债监控实例正在运行") from exc
        yield
    finally:
        if os.name == "nt":
            try:
                handle.seek(0)
                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            except OSError:
                pass
        else:
            import fcntl

            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        handle.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="可转债低频监控服务")
    parser.add_argument("--workbook", type=Path, help="监控 Excel 路径")
    commands = parser.add_subparsers(dest="command")
    init = commands.add_parser("init", help="创建 Excel 面板")
    init.add_argument("--legacy", type=Path, help="从旧 kzz 工作簿导入代码")
    commands.add_parser("run", help="显示控制台并在托盘后台运行")
    commands.add_parser("headless", help="无控制台持续运行（高级用途）")
    commands.add_parser("once", help="忽略交易时间，立即完整更新一轮")
    import_adq = commands.add_parser("import-adq", help="立即导入安道全评级和三段线")
    import_adq.add_argument("source", type=Path)
    import_adq.add_argument("--sheet", default="")
    commands.add_parser("test-notify", help="测试桌面和邮件提醒")
    return parser


def main() -> None:
    if "--version-probe" in sys.argv:
        import sqlite3
        import ssl
        from . import __version__

        (application_dir() / "exe-self-test.txt").write_text(
            f"ok\nversion={__version__}\nssl={ssl.OPENSSL_VERSION}\nsqlite={sqlite3.sqlite_version}\n",
            encoding="utf-8",
        )
        return
    if "--data-probe" in sys.argv:
        from .clock import china_now
        from .provider import AkShareProvider

        provider = AkShareProvider()
        trade_date = provider.is_trade_date(china_now().date())
        (application_dir() / "data-self-test.txt").write_text(
            f"ok\ntrade_date={trade_date}\n",
            encoding="utf-8",
        )
        return
    args = build_parser().parse_args()
    if args.command is None:
        args.command = "run"
    base = application_dir()
    base.mkdir(parents=True, exist_ok=True)
    setup_logging(base)
    workbook = (args.workbook or base / "可转债监控.xlsx").resolve()
    state_path = base / "data" / "monitor.db"

    if args.command == "init":
        legacy = args.legacy or (base.parent / "kzz_vba.xlsm")
        codes = legacy_codes(legacy)
        if workbook.exists():
            raise SystemExit(f"文件已经存在，不会覆盖：{workbook}")
        create_workbook(workbook, codes or None)
        print(f"已创建：{workbook}；导入旧代码 {len(codes)} 条")
        return
    if not workbook.exists():
        template = bundled_resource("可转债监控.xlsx")
        if template:
            ensure_workbook(workbook)
        else:
            create_workbook(workbook, legacy_codes(base.parent / "kzz_vba.xlsm") or None)
        logging.info("首次运行，已创建 Excel 面板：%s", workbook)
    try:
        migrate_workbook(workbook)
    except PermissionError:
        logging.warning("Excel 正在打开，本次暂不执行工作簿升级；关闭 Excel 后重启即可")

    if args.command == "import-adq":
        count = ExcelStore(workbook).import_adq(args.source.resolve(), args.sheet)
        print(f"已导入 {count} 条评级和三段线")
        return
    if args.command == "test-notify":
        settings, _ = load_configuration(workbook)
        Notifier(settings, state_path.parent / "smtp_secret.bin").send(
            "提醒测试", "桌面/邮件提醒配置有效。"
        )
        return

    try:
        with single_instance(base):
            if args.command == "run":
                from .gui import run_gui

                run_gui(workbook, state_path)
                return
            service = MonitorService(workbook, state_path)
            if args.command == "once":
                service.run_once()
            else:
                service.run()
    except RuntimeError as exc:
        logging.warning("%s", exc)
        show_already_running("KzzMonitor 已经在运行。请从系统托盘或 macOS 菜单栏打开现有控制台。")


if __name__ == "__main__":
    main()
