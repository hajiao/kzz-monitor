from datetime import datetime, time
import hashlib
import json
from pathlib import Path
import threading
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from kzz_monitor.config import AppSettings, BondConfig, normalize_code
from kzz_monitor.config import create_workbook, load_configuration, migrate_workbook
from kzz_monitor.excel_store import ExcelStore
from kzz_monitor.models import Evaluation, Quote, Trend
from kzz_monitor.monitor_view import classify_bond_row
from kzz_monitor.provider import exchange_symbol
from kzz_monitor.service import AlertEngine, MonitorService
from kzz_monitor.storage import StateStore
from kzz_monitor.updater import check_for_update
from kzz_monitor.updater import _launch_windows_replacer


def quote(price: float, minute: int) -> Quote:
    return Quote("127015", "希望转债", price, datetime(2026, 9, 3, 10, minute))


def test_sell_alert_requires_arming_downtrend_and_drawdown(tmp_path):
    state = StateStore(tmp_path / "state.db")
    engine = AlertEngine(state)
    config = BondConfig("127015", sell_trigger_price=130, sell_drawdown_pct=5, trend_window=2, trend_epsilon=0.1)

    assert not engine.evaluate(config, quote(140, 0), 145).sell_alert
    result = engine.evaluate(config, quote(132, 1), 145)
    assert result.trend == Trend.DOWN
    assert result.sell_alert
    assert result.drawdown_pct == pytest.approx(5.714, rel=1e-3)
    assert "卖出观察" in result.alert_messages[0]

    result = engine.evaluate(config, quote(130, 2), 145)
    assert result.sell_alert
    assert not result.alert_messages  # 同一轮下跌只提醒一次
    state.close()


def test_position_zone_only_alerts_when_moving_deeper(tmp_path):
    state = StateStore(tmp_path / "state.db")
    engine = AlertEngine(state)
    config = BondConfig("127015", build_line=120, add_line=115, heavy_line=110)

    assert engine.evaluate(config, quote(119, 0), 130).alert_messages == ["进入建仓：现价 119.00"]
    assert engine.evaluate(config, quote(114, 1), 130).alert_messages == ["进入加仓：现价 114.00"]
    assert engine.evaluate(config, quote(116, 2), 130).alert_messages == []
    state.close()


def test_market_phases():
    settings = AppSettings()
    assert MonitorService._market_phase(time(9, 0), settings) == "before"
    assert MonitorService._market_phase(time(9, 30), settings) == "open"
    assert MonitorService._market_phase(time(12, 0), settings) == "lunch"
    assert MonitorService._market_phase(time(14, 59), settings) == "open"
    assert MonitorService._market_phase(time(15, 0), settings) == "closed"
    assert MonitorService._phase_wait_message("before", settings) == "开盘前，等待 09:30 开市"
    assert MonitorService._phase_wait_message("lunch", settings) == "午间休市，等待 13:00 恢复交易"
    assert MonitorService._phase_wait_message("closed", settings) == "今日已收盘，等待下一交易日"


def test_cycle_delay_is_start_to_start_and_never_negative():
    assert MonitorService._cycle_delay_seconds(60, 15 * 60) == 45 * 60
    assert MonitorService._cycle_delay_seconds(60, 60 * 60) == 0
    assert MonitorService._cycle_delay_seconds(60, 70 * 60) == 0


def test_wait_until_next_day_preserves_timezone(tmp_path):
    class ProviderStub:
        pass

    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    current = datetime(2026, 9, 3, 19, 9, 58, tzinfo=ZoneInfo("Asia/Shanghai"))
    service = MonitorService(
        workbook,
        tmp_path / "monitor.db",
        provider=ProviderStub(),  # type: ignore[arg-type]
        now=lambda: current,
    )
    waits: list[float] = []
    service._wait = lambda seconds: waits.append(seconds) or False  # type: ignore[method-assign]
    service._wait_until_next_day()
    assert waits == [3600]
    service.state.close()


def test_last_complete_cycle_is_persisted_by_date(tmp_path):
    class ProviderStub:
        pass

    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    current = datetime(2026, 9, 4, 12, 0, tzinfo=ZoneInfo("Asia/Shanghai"))
    service = MonitorService(
        workbook,
        tmp_path / "monitor.db",
        provider=ProviderStub(),  # type: ignore[arg-type]
        now=lambda: current,
    )
    assert not service._last_cycle_is_today(current.date())
    assert service._needs_lunch_catchup("lunch", current.date())
    assert not service._needs_lunch_catchup("before", current.date())
    service._mark_cycle_completed()
    assert service._last_cycle_is_today(current.date())
    assert not service._needs_lunch_catchup("lunch", current.date())
    assert not service._last_cycle_is_today(datetime(2026, 9, 5).date())
    service.state.close()


def test_code_and_exchange_normalization():
    assert normalize_code("SZ127015") == "127015"
    assert normalize_code(110059.0) == "110059"
    assert exchange_symbol("113056") == "sh113056"
    assert exchange_symbol("127015") == "sz127015"


def test_excel_accepts_timezone_aware_quote(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["127015"])
    current = Quote("127015", "希望转债", 123.45, datetime(2026, 9, 3, 10, 0, tzinfo=ZoneInfo("Asia/Shanghai")))
    evaluation = Evaluation(Trend.FLAT, 140, 130, 5, "观察", False, [])
    ExcelStore(workbook).update_result(current, evaluation)


def test_excel_marks_unavailable_bond_without_error(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["127015"])
    ExcelStore(workbook).mark_unavailable(
        "127015",
        datetime(2026, 9, 3, 10, 0, tzinfo=ZoneInfo("Asia/Shanghai")),
        "最后交易日 2025-12-26，可能已退市",
    )


def test_state_store_can_move_from_gui_thread_to_worker(tmp_path):
    state = StateStore(tmp_path / "threaded.db")
    errors: list[BaseException] = []

    def worker() -> None:
        try:
            state.set_daily_high("113043", datetime(2026, 9, 3).date(), 145.5)
            assert state.get_daily_high("113043", datetime(2026, 9, 3).date()) == 145.5
            state.add_price("113043", datetime(2026, 9, 3, 10, 0), 120.5)
        except BaseException as exc:
            errors.append(exc)

    thread = threading.Thread(target=worker)
    thread.start()
    thread.join()
    assert errors == []
    assert state.recent_prices("113043", 1) == [120.5]
    state.close()


def test_monitor_service_retains_state_path_for_notifier(tmp_path):
    class ProviderStub:
        pass

    workbook = tmp_path / "monitor.xlsx"
    database = tmp_path / "data" / "monitor.db"
    create_workbook(workbook, ["113043"])
    service = MonitorService(workbook, database, provider=ProviderStub())  # type: ignore[arg-type]
    assert service.state_path == database
    assert service.state_path.parent / "smtp_secret.bin" == tmp_path / "data" / "smtp_secret.bin"
    service.state.close()


def test_workbook_migration_preserves_existing_settings(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    migrate_workbook(workbook)
    settings, bonds = load_configuration(workbook)
    assert settings.cycle_interval_minutes == 60
    assert bonds[0].code == "113043"


def test_local_update_manifest_resolves_relative_package(tmp_path):
    package = tmp_path / "KzzMonitor-update.zip"
    package.write_bytes(b"example-update")
    manifest = tmp_path / "update-manifest.json"
    manifest.write_text(json.dumps({
        "version": "9.0.0",
        "notes": "test",
        "artifacts": {key: {
                "url": package.name,
                "sha256": hashlib.sha256(package.read_bytes()).hexdigest(),
            } for key in ("windows-x64", "macos-arm64", "macos-x64")},
    }), encoding="utf-8")
    info = check_for_update(str(manifest))
    assert info is not None
    assert Path(info.url) == package.resolve()


def test_excel_pending_queue_persists_and_flushes(tmp_path):
    class ProviderStub:
        pass

    workbook = tmp_path / "monitor.xlsx"
    database = tmp_path / "data" / "monitor.db"
    create_workbook(workbook, ["113043"])
    service = MonitorService(workbook, database, provider=ProviderStub())  # type: ignore[arg-type]
    current = Quote("113043", "财通转债", 121.5, datetime(2026, 9, 3, 10, 0))
    evaluation = Evaluation(Trend.DOWN, 140, 130, 6.5, "观察", True, ["卖出观察"])
    service._queue_result_write(current, evaluation)
    service._queue_alert_write(current, "卖出", "卖出观察")
    assert service.state.pending_excel_write_count() == 2
    service.state.close()

    restarted = MonitorService(workbook, database, provider=ProviderStub())  # type: ignore[arg-type]
    assert restarted.state.pending_excel_write_count() == 2
    restarted._flush_pending_excel_writes()
    assert restarted.state.pending_excel_write_count() == 0
    restarted.state.close()


def test_monitor_crud_prevents_and_merges_duplicates(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    store = ExcelStore(workbook)
    values = {
        "启用": True, "转债代码": "SH113043", "名称": "财通转债",
        "卖出观察价": 135, "回撤提醒%": 6, "趋势窗口": 4,
        "趋势最小跌幅": 0.2, "建仓线": 120, "加仓线": 115,
        "重仓线": 110, "当周评价": "测试",
    }
    code, created = store.upsert_bond(values)
    assert (code, created) == ("113043", False)
    assert len(store.list_bonds()) == 1

    wb = load_workbook(workbook)
    sheet = wb["监控列表"]
    sheet.append([True, "113043", "重复行", 130, 5, 3, 0.1])
    wb.save(workbook)
    wb.close()
    _, loaded = load_configuration(workbook)
    assert [bond.code for bond in loaded] == ["113043"]
    assert store.deduplicate_bonds() == {"113043": 1}
    assert len(store.list_bonds()) == 1
    assert store.delete_bond("113043")
    assert store.list_bonds() == []


def test_monitor_crud_rejects_invalid_lines(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    with pytest.raises(ValueError, match="建仓线"):
        ExcelStore(workbook).upsert_bond({
            "启用": True, "转债代码": "113043", "卖出观察价": 130,
            "回撤提醒%": 5, "趋势窗口": 3, "趋势最小跌幅": 0.1,
            "建仓线": 110, "加仓线": 115, "重仓线": 100,
        })


def test_blank_enabled_cell_defaults_to_enabled(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["127049"])
    wb = load_workbook(workbook)
    wb["监控列表"]["A2"] = None
    wb.save(workbook)
    wb.close()
    _, bonds = load_configuration(workbook)
    assert [bond.code for bond in bonds] == ["127049"]


def test_update_schedule_defaults_and_minimum(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    settings, _ = load_configuration(workbook)
    assert settings.check_updates_on_startup is True
    assert settings.update_check_interval_hours == 24
    assert settings.auto_install_updates is False


def test_monitor_visual_alert_priority_and_contrast_states():
    base = {
        "当前价": 120, "监控峰值": 140, "卖出观察价": 130,
        "回撤%": 14.2, "回撤提醒%": 5, "趋势": "下降",
        "仓位区域": "重仓", "强赎状态": "已公告强赎",
    }
    assert classify_bond_row(base).key == "sell"
    assert classify_bond_row({**base, "趋势": "震荡"}).key == "heavy"
    assert classify_bond_row({**base, "趋势": "震荡", "仓位区域": "加仓"}).key == "add"
    assert classify_bond_row({**base, "趋势": "震荡", "仓位区域": "建仓"}).key == "build"
    assert classify_bond_row({**base, "趋势": "震荡", "仓位区域": "观察"}).key == "redeem"
    assert classify_bond_row({**base, "趋势": "无实时行情"}).key == "unavailable"
    assert classify_bond_row({"趋势": "震荡", "强赎状态": "未开始"}).key == "normal"
    assert classify_bond_row({"趋势": "震荡"}, sell_latched=True).key == "sell"


def test_windows_update_replacer_has_retry_restart_and_log(tmp_path, monkeypatch):
    calls: list[list[str]] = []
    monkeypatch.setattr(
        "kzz_monitor.updater.subprocess.Popen",
        lambda command, **_kwargs: calls.append(command),
    )
    staging = tmp_path / "staging"
    payload = staging / "payload"
    target = tmp_path / "app"
    payload.mkdir(parents=True)
    target.mkdir()
    (payload / "KzzMonitor.exe").write_bytes(b"new")
    _launch_windows_replacer(staging, payload, target)
    script = (staging / "install-update.ps1").read_text(encoding="utf-8-sig")
    assert "for ($i=1; $i -le 20; $i++)" in script
    assert "for ($i=1; $i -le 3; $i++)" in script
    assert "重启成功" in script
    assert "update.log" in " ".join(calls[0])
