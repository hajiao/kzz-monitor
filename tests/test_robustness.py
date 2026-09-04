from __future__ import annotations

import json
import os
import queue
import threading
import time
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import kzz_monitor.excel_store as excel_module
import kzz_monitor.updater as updater
from kzz_monitor.config import AppSettings, create_workbook, load_configuration, update_settings
from kzz_monitor.excel_store import ExcelStore
from kzz_monitor.gui import MonitorApp
from kzz_monitor.models import BondConfig, Quote
from kzz_monitor.provider import AkShareProvider, _number
from kzz_monitor.service import MonitorService
from kzz_monitor.storage import StateStore


class ProviderStub:
    def __init__(self) -> None:
        self.refresh_count = 0

    def refresh_spot(self, now=None):
        self.refresh_count += 1
        timestamp = now or datetime(2026, 9, 4, 10, 0)
        return {
            "113043": Quote("113043", "财通转债", 120, timestamp),
            "113056": Quote("113056", "重银转债", 125, timestamp),
        }


def test_one_spot_request_per_complete_cycle(tmp_path, monkeypatch):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043", "113056"])
    provider = ProviderStub()
    service = MonitorService(workbook, tmp_path / "state.db", provider=provider)  # type: ignore[arg-type]
    processed: list[tuple[str, dict[str, Quote]]] = []
    monkeypatch.setattr(service, "_process", lambda config, settings, quotes=None: processed.append((config.code, quotes)))
    assert service._run_complete_cycle(AppSettings(), [BondConfig("113043"), BondConfig("113056")], False)
    assert provider.refresh_count == 1
    assert [code for code, _ in processed] == ["113043", "113056"]
    assert processed[0][1] is processed[1][1]
    service.state.close()


def test_after_hours_cycle_resumes_after_restart_without_repeating_completed_bonds(tmp_path, monkeypatch):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043", "113056", "127049"])
    database = tmp_path / "state.db"
    settings = AppSettings(poll_interval_seconds=10)
    bonds = [BondConfig("113043"), BondConfig("113056"), BondConfig("127049")]

    first_provider = ProviderStub()
    first_provider.refresh_spot = lambda now=None: {}  # type: ignore[method-assign]
    first = MonitorService(workbook, database, provider=first_provider)  # type: ignore[arg-type]
    first_processed: list[str] = []
    monkeypatch.setattr(first, "_process", lambda config, _settings, quotes=None: first_processed.append(config.code))
    monkeypatch.setattr(first, "_wait", lambda _seconds: True)
    assert not first._run_resumable_final_cycle(settings, bonds, datetime(2026, 9, 4).date())
    assert first_processed == ["113043"]
    assert json.loads(first.state.get_value("final_cycle_progress:2026-09-04")) == ["113043"]
    first.state.close()

    second_provider = ProviderStub()
    second_provider.refresh_spot = lambda now=None: {}  # type: ignore[method-assign]
    second = MonitorService(workbook, database, provider=second_provider)  # type: ignore[arg-type]
    second_processed: list[str] = []
    monkeypatch.setattr(second, "_process", lambda config, _settings, quotes=None: second_processed.append(config.code))
    monkeypatch.setattr(second, "_wait", lambda _seconds: False)
    assert second._run_resumable_final_cycle(settings, bonds, datetime(2026, 9, 4).date())
    assert second_processed == ["113056", "127049"]
    assert second.state.get_value("final_cycle_progress:2026-09-04") == ""
    second.state.close()


def test_corrupt_after_hours_checkpoint_recovers_from_empty_progress(tmp_path, monkeypatch):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    service = MonitorService(workbook, tmp_path / "state.db", provider=ProviderStub())  # type: ignore[arg-type]
    service.state.set_value("final_cycle_progress:2026-09-04", "not-json")
    processed: list[str] = []
    monkeypatch.setattr(service, "_process", lambda config, _settings, quotes=None: processed.append(config.code))
    assert service._run_resumable_final_cycle(
        AppSettings(), [BondConfig("113043")], datetime(2026, 9, 4).date()
    )
    assert processed == ["113043"]
    service.state.close()


def test_poison_pending_write_does_not_block_valid_item(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    database = tmp_path / "state.db"
    service = MonitorService(workbook, database, provider=ProviderStub())  # type: ignore[arg-type]
    now = datetime(2026, 9, 4, 10, 0)
    service.state.queue_excel_write("bad:first", "result", "{not-json", now)
    quote = Quote("113043", "财通转债", 120.5, now)
    from kzz_monitor.models import Evaluation, Trend
    service._queue_result_write(quote, Evaluation(Trend.FLAT, 130, 125, 3.6, "观察", False, []))
    service._flush_pending_excel_writes()
    assert service.state.pending_excel_write_count() == 0
    assert service.state.failed_excel_write_count() == 1
    rows = service.excel.list_bonds()
    assert rows[0]["当前价"] == 120.5
    service.state.close()


def test_state_store_close_race_has_only_explicit_closed_errors(tmp_path):
    state = StateStore(tmp_path / "state.db")
    start = threading.Event()
    errors: list[BaseException] = []

    def worker(index: int) -> None:
        start.wait()
        for count in range(500):
            try:
                state.set_value(f"{index}:{count}", str(count))
                state.get_value(f"{index}:{count}")
            except BaseException as exc:
                errors.append(exc)
                return

    threads = [threading.Thread(target=worker, args=(index,)) for index in range(6)]
    for thread in threads:
        thread.start()
    start.set()
    time.sleep(0.01)
    state.close()
    for thread in threads:
        thread.join(timeout=5)
        assert not thread.is_alive(), "database worker deadlocked"
    assert all(isinstance(error, RuntimeError) and "已经关闭" in str(error) for error in errors)


def test_concurrent_excel_upserts_are_serialized(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    store = ExcelStore(workbook)
    errors: list[BaseException] = []

    def add(code: str) -> None:
        try:
            store.upsert_bond({
                "启用": True, "转债代码": code, "卖出观察价": 130,
                "回撤提醒%": 5, "趋势窗口": 3, "趋势最小跌幅": 0.1,
            })
        except BaseException as exc:
            errors.append(exc)

    codes = [f"123{index:03d}" for index in range(20)]
    threads = [threading.Thread(target=add, args=(code,)) for code in codes]
    for thread in threads:
        thread.start()
    deadline = time.monotonic() + 30
    for thread in threads:
        thread.join(timeout=max(0, deadline - time.monotonic()))
        assert not thread.is_alive(), "Excel writer deadlocked"
    assert errors == []
    assert {row["转债代码"] for row in store.list_bonds()} == {"113043", *codes}
    load_workbook(workbook).close()


def test_excel_failed_replace_leaves_original_valid_and_no_temp(tmp_path, monkeypatch):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    store = ExcelStore(workbook)
    monkeypatch.setattr(
        excel_module,
        "replace_with_retry",
        lambda *_: (_ for _ in ()).throw(PermissionError("locked")),
    )
    with pytest.raises(PermissionError, match="请关闭"):
        store.upsert_bond({
            "启用": True, "转债代码": "113056", "卖出观察价": 130,
            "回撤提醒%": 5, "趋势窗口": 3, "趋势最小跌幅": 0.1,
        })
    assert [bond.code for bond in load_configuration(workbook)[1]] == ["113043"]
    assert list(tmp_path.glob(".*.writing.xlsx")) == []


def test_update_settings_malformed_workbook_does_not_mask_original_error(tmp_path):
    workbook = tmp_path / "broken.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "错误表名"
    wb.save(workbook)
    wb.close()
    with pytest.raises(KeyError):
        update_settings(workbook, {"轮询间隔秒": 60})
    assert list(tmp_path.glob(".*.settings.xlsx")) == []


def test_upsert_automatically_removes_existing_duplicates(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    wb = load_workbook(workbook)
    wb["监控列表"].append([True, "113043", "重复名称", 130, 5, 3, 0.1])
    wb.save(workbook)
    wb.close()
    ExcelStore(workbook).upsert_bond({
        "启用": True, "转债代码": "113043", "名称": "财通转债",
        "卖出观察价": 135, "回撤提醒%": 6, "趋势窗口": 4, "趋势最小跌幅": 0.2,
    })
    rows = ExcelStore(workbook).list_bonds()
    assert len(rows) == 1
    assert rows[0]["名称"] == "财通转债"
    assert rows[0]["卖出观察价"] == 135


def test_provider_rejects_non_finite_numbers():
    assert _number(float("nan")) is None
    assert _number("inf") is None
    assert _number("-inf") is None
    assert _number("123.45") == 123.45


def test_provider_sorts_history_before_last_trade():
    provider = object.__new__(AkShareProvider)
    frame = pd.DataFrame({
        "date": ["2026-01-03", "2026-01-01", "2026-01-02"],
        "close": [103, 101, 102], "high": [104, 102, 103],
    })
    provider.history = lambda code: frame  # type: ignore[method-assign]
    assert provider.last_trade("113043") == (datetime(2026, 1, 3).date(), 103)
    assert provider.one_year_high("113043", datetime(2026, 1, 3).date()) == 104


def test_provider_skips_nan_and_infinite_spot_prices():
    provider = object.__new__(AkShareProvider)
    frame = pd.DataFrame({
        "代码": ["113043", "113056", "127049"],
        "名称": ["财通转债", "重银转债", "希望转2"],
        "现价": [float("nan"), float("inf"), 111.4],
    })
    quotes = provider._parse_spot(frame, datetime(2026, 9, 4, 10, 0))
    assert list(quotes) == ["127049"]


def test_service_wait_is_interruptible_without_deadlock(tmp_path):
    workbook = tmp_path / "monitor.xlsx"
    create_workbook(workbook, ["113043"])
    service = MonitorService(workbook, tmp_path / "state.db", provider=ProviderStub())  # type: ignore[arg-type]
    result: list[bool] = []
    thread = threading.Thread(target=lambda: result.append(service._wait(30)))
    thread.start()
    time.sleep(0.02)
    service.request_force_cycle()
    thread.join(timeout=2)
    assert not thread.is_alive()
    assert result == [True]
    service.state.close()


def test_ui_action_queue_executes_callbacks_on_drain():
    class FakeRoot:
        def __init__(self):
            self.after_calls = []

        def after(self, delay, callback, *args):
            self.after_calls.append((delay, callback, args))
            return "after-id"

    app = object.__new__(MonitorApp)
    app.quitting = False
    app.root = FakeRoot()
    app.ui_actions = queue.Queue()
    received = []
    app._post_ui(received.append, "from-worker")
    app._drain_ui_actions()
    assert received == ["from-worker"]
    assert app.root.after_calls[0][0] == 100


def write_zip(path: Path, members: list[tuple[zipfile.ZipInfo | str, bytes]]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in members:
            archive.writestr(name, content)


@pytest.mark.parametrize("name", ["../evil.exe", "/absolute.exe", "C:/evil.exe", "folder/../../evil.exe"])
def test_updater_rejects_path_traversal(tmp_path, name):
    archive = tmp_path / "bad.zip"
    write_zip(archive, [(name, b"bad")])
    with pytest.raises(ValueError, match="不安全路径"):
        updater._validate_archive(archive)


def test_updater_rejects_duplicate_casefolded_targets(tmp_path):
    archive = tmp_path / "bad.zip"
    write_zip(archive, [("A.txt", b"one"), ("a.txt", b"two")])
    with pytest.raises(ValueError, match="重复目标"):
        updater._validate_archive(archive)


def test_updater_rejects_symlink(tmp_path):
    archive = tmp_path / "bad.zip"
    info = zipfile.ZipInfo("link")
    info.create_system = 3
    info.external_attr = (0o120777 << 16)
    write_zip(archive, [(info, b"target")])
    with pytest.raises(ValueError, match="符号链接"):
        updater._validate_archive(archive)


def test_updater_rejects_zip_bomb_ratio(tmp_path, monkeypatch):
    archive = tmp_path / "bad.zip"
    write_zip(archive, [("huge.txt", b"0" * 100_000)])
    monkeypatch.setattr(updater, "MAX_COMPRESSION_RATIO", 5)
    with pytest.raises(ValueError, match="压缩比"):
        updater._validate_archive(archive)


def test_updater_rejects_http_and_oversized_local_manifest(tmp_path):
    with pytest.raises(ValueError, match="HTTPS"):
        updater._read_location("http://example.com/update.json")
    manifest = tmp_path / "manifest.json"
    manifest.write_bytes(b"x" * 100)
    with pytest.raises(ValueError, match="大小"):
        updater._read_location(str(manifest), max_size=10)
